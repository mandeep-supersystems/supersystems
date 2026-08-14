import uuid
import io
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import verify_jwt_in_request, get_jwt_identity
from extensions import db

asset_bp = Blueprint("assets", __name__)

TENANT_ID = "b424df0e-f766-4e94-b3fd-05777e158958"
TID_COND = "(tenant_id = :tid OR tenant_id = 'TEST' OR tenant_id = 'b424df0e-f766-4e94-b3fd-05777e158958' OR tenant_id = '' OR tenant_id IS NULL)"


def _get_tenant():
    try:
        verify_jwt_in_request(optional=True)
        identity = get_jwt_identity()
        if isinstance(identity, dict):
            return identity.get("tenant_id", "TEST")
        if isinstance(identity, str):
            import json
            try:
                return json.loads(identity).get("tenant_id", "TEST")
            except Exception:
                pass
    except Exception:
        pass
    return "TEST"


# ── SERIES ──────────────────────────────────────────────

@asset_bp.route("/series", methods=["GET"])
def list_series():
    tid = _get_tenant()
    rows = db.session.execute(db.text(
        f"SELECT id, series_number, category, description, "
        f"(SELECT COUNT(*) FROM asset.items i WHERE i.series_number = s.series_number AND i.is_deleted = false) as item_count "
        f"FROM asset.series s WHERE s.is_deleted = false AND {TID_COND} ORDER BY series_number ASC"
    ), {"tid": tid}).fetchall()
    return jsonify({"success": True, "data": [
        {"id": r[0], "series_number": r[1], "category": r[2], "description": r[3] or "", "item_count": int(r[4] or 0)}
        for r in rows
    ]})


@asset_bp.route("/series", methods=["POST"])
def create_series():
    tid = _get_tenant()
    data = request.get_json() or {}
    series_number = data.get("series_number")
    category = (data.get("category") or "").strip()
    if not series_number or not category:
        return jsonify({"success": False, "message": "series_number and category are required"}), 400
    existing = db.session.execute(db.text(
        "SELECT id FROM asset.series WHERE series_number = :sn AND is_deleted = false"
    ), {"sn": series_number}).first()
    if existing:
        return jsonify({"success": False, "message": f"Series {series_number} already exists"}), 409
    sid = str(uuid.uuid4())
    db.session.execute(db.text(
        "INSERT INTO asset.series (id, series_number, category, description, tenant_id) "
        "VALUES (:id, :sn, :cat, :desc, :tid)"
    ), {"id": sid, "sn": series_number, "cat": category, "desc": data.get("description", ""), "tid": TENANT_ID})
    db.session.commit()
    return jsonify({"success": True, "message": f"Series {series_number} created", "id": sid})


# ── ITEMS ────────────────────────────────────────────────

@asset_bp.route("/register", methods=["GET"])
def list_assets():
    tid = _get_tenant()
    series_filter = request.args.get("series", "").strip()
    search = request.args.get("search", "").strip()

    where = f"WHERE i.is_deleted = false AND {TID_COND.replace('tenant_id', 'i.tenant_id')}"
    params = {"tid": tid}
    if series_filter:
        where += " AND i.series_number = :sn"
        params["sn"] = series_filter
    if search:
        where += " AND (i.asset_number ILIKE :s OR i.category ILIKE :s OR i.description ILIKE :s OR i.make ILIKE :s)"
        params["s"] = f"%{search}%"

    rows = db.session.execute(db.text(
        f"SELECT i.id, i.series_number, i.asset_number, i.category, i.description, i.qty, i.make, i.status, i.location "
        f"FROM asset.items i {where} ORDER BY i.series_number ASC, i.asset_number ASC"
    ), params).fetchall()

    return jsonify({"success": True, "data": [
        {"id": r[0], "series_number": r[1], "asset_number": r[2], "category": r[3],
         "description": r[4] or "", "qty": int(r[5] or 1), "make": r[6] or "",
         "status": r[7] or "active", "location": r[8] or ""}
        for r in rows
    ]})


@asset_bp.route("/register", methods=["POST"])
def add_asset():
    tid = _get_tenant()
    data = request.get_json() or {}
    series_number = data.get("series_number")
    category = (data.get("category") or "").strip()
    if not series_number or not category:
        return jsonify({"success": False, "message": "series_number and category are required"}), 400

    # Get series, auto-create if missing
    series_row = db.session.execute(db.text(
        "SELECT id FROM asset.series WHERE series_number = :sn AND is_deleted = false"
    ), {"sn": series_number}).first()
    if not series_row:
        sid = str(uuid.uuid4())
        db.session.execute(db.text(
            "INSERT INTO asset.series (id, series_number, category, description, tenant_id) "
            "VALUES (:id, :sn, :cat, :desc, :tid)"
        ), {"id": sid, "sn": series_number, "cat": category, "desc": "", "tid": TENANT_ID})
        series_id = sid
    else:
        series_id = series_row[0]

    # Auto-generate asset_number: SERIES.NNNN
    last = db.session.execute(db.text(
        "SELECT asset_number FROM asset.items WHERE series_number = :sn AND is_deleted = false "
        "ORDER BY asset_number DESC LIMIT 1"
    ), {"sn": series_number}).scalar()
    if last:
        try:
            seq = int(last.split(".")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    asset_number = f"{series_number}.{str(seq).zfill(4)}"

    iid = str(uuid.uuid4())
    db.session.execute(db.text(
        "INSERT INTO asset.items (id, series_id, series_number, asset_number, category, description, qty, make, status, location, tenant_id) "
        "VALUES (:id, :sid, :sn, :an, :cat, :desc, :qty, :make, :status, :loc, :tid)"
    ), {
        "id": iid, "sid": series_id, "sn": series_number, "an": asset_number,
        "cat": category, "desc": data.get("description", ""),
        "qty": int(data.get("qty", 1)), "make": data.get("make", ""),
        "status": data.get("status", "active"), "loc": data.get("location", ""),
        "tid": TENANT_ID
    })
    db.session.commit()
    return jsonify({"success": True, "message": f"Asset {asset_number} added", "asset_number": asset_number})


@asset_bp.route("/register/<item_id>", methods=["PUT"])
def update_asset(item_id):
    data = request.get_json() or {}
    fields, params = [], {"id": item_id}
    for col in ["description", "qty", "make", "status", "location"]:
        if col in data:
            fields.append(f"{col}=:{col}")
            params[col] = data[col]
    if not fields:
        return jsonify({"success": False, "message": "Nothing to update"}), 400
    db.session.execute(db.text(
        f"UPDATE asset.items SET {', '.join(fields)} WHERE id=:id"
    ), params)
    db.session.commit()
    return jsonify({"success": True, "message": "Asset updated"})


@asset_bp.route("/register/<item_id>", methods=["DELETE"])
def delete_asset(item_id):
    db.session.execute(db.text(
        "UPDATE asset.items SET is_deleted=true WHERE id=:id"
    ), {"id": item_id})
    db.session.commit()
    return jsonify({"success": True, "message": "Asset deleted"})


# ── IMPORT from Excel ────────────────────────────────────

@asset_bp.route("/import", methods=["POST"])
def import_assets():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
    f = request.files["file"]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        inserted = 0
        skipped = 0
        series_cache = {}   # series_number -> series_id
        seq_cache = {}      # series_number -> next sequence int

        def _ensure_series(sn, category):
            if sn in series_cache:
                return series_cache[sn]
            sr = db.session.execute(db.text(
                "SELECT id FROM asset.series WHERE series_number=:sn AND is_deleted=false"
            ), {"sn": sn}).first()
            if not sr:
                sid = str(uuid.uuid4())
                db.session.execute(db.text(
                    "INSERT INTO asset.series (id, series_number, category, description, tenant_id) "
                    "VALUES (:id, :sn, :cat, '', :tid)"
                ), {"id": sid, "sn": sn, "cat": category, "tid": TENANT_ID})
                series_cache[sn] = sid
            else:
                series_cache[sn] = sr[0]
            return series_cache[sn]

        def _next_asset_number(sn):
            if sn not in seq_cache:
                last = db.session.execute(db.text(
                    "SELECT asset_number FROM asset.items WHERE series_number=:sn AND is_deleted=false "
                    "ORDER BY asset_number DESC LIMIT 1"
                ), {"sn": sn}).scalar()
                if last:
                    try:
                        seq_cache[sn] = int(str(last).split(".")[-1])
                    except Exception:
                        seq_cache[sn] = 0
                else:
                    seq_cache[sn] = 0
            seq_cache[sn] += 1
            return f"{sn}.{str(seq_cache[sn]).zfill(4)}"

        current_series = None
        current_category = ""

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            raw_series   = row[0]
            raw_asset_no = row[1]
            category     = str(row[2]).strip() if row[2] else current_category
            description  = str(row[3]).strip() if row[3] else ""
            qty          = int(row[4]) if row[4] is not None else 1
            make         = str(row[5]).strip() if row[5] else ""

            # Update running series
            if raw_series is not None:
                try:
                    current_series = int(float(raw_series))
                except Exception:
                    continue
            if category:
                current_category = category

            if current_series is None or not current_category:
                skipped += 1
                continue

            # Resolve asset_number
            if raw_asset_no is not None:
                # Excel stores as float e.g. 5252.0001 — format to SERIES.NNNN
                try:
                    parts = f"{float(raw_asset_no):.4f}".split(".")
                    asset_number = f"{int(parts[0])}.{parts[1]}"
                except Exception:
                    asset_number = str(raw_asset_no).strip()
            else:
                asset_number = _next_asset_number(current_series)

            # Skip duplicates
            exists = db.session.execute(db.text(
                "SELECT id FROM asset.items WHERE asset_number=:an AND is_deleted=false"
            ), {"an": asset_number}).first()
            if exists:
                skipped += 1
                continue

            series_id = _ensure_series(current_series, current_category)

            db.session.execute(db.text(
                "INSERT INTO asset.items (id, series_id, series_number, asset_number, category, description, qty, make, tenant_id) "
                "VALUES (:id, :sid, :sn, :an, :cat, :desc, :qty, :make, :tid)"
            ), {
                "id": str(uuid.uuid4()), "sid": series_id,
                "sn": current_series, "an": asset_number,
                "cat": current_category, "desc": description,
                "qty": qty, "make": make, "tid": TENANT_ID
            })
            inserted += 1

        db.session.commit()
        msg = f"{inserted} assets imported"
        if skipped:
            msg += f", {skipped} skipped (duplicates or missing data)"
        return jsonify({"success": True, "message": msg})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


# ── OVERVIEW STATS ─────────────────────────────────────────

@asset_bp.route("/overview-stats", methods=["GET"])
def overview_stats():
    try:
        total = db.session.execute(db.text(
            "SELECT COUNT(*) FROM asset.items WHERE is_deleted=false"
        )).scalar() or 0
        series = db.session.execute(db.text(
            "SELECT COUNT(*) FROM asset.series WHERE is_deleted=false"
        )).scalar() or 0
        active = db.session.execute(db.text(
            "SELECT COUNT(*) FROM asset.items WHERE is_deleted=false AND status='active'"
        )).scalar() or 0
        disposed = db.session.execute(db.text(
            "SELECT COUNT(*) FROM asset.items WHERE is_deleted=false AND status='disposed'"
        )).scalar() or 0
        under_repair = db.session.execute(db.text(
            "SELECT COUNT(*) FROM asset.items WHERE is_deleted=false AND status='under_repair'"
        )).scalar() or 0
        total_qty = db.session.execute(db.text(
            "SELECT COALESCE(SUM(qty),0) FROM asset.items WHERE is_deleted=false"
        )).scalar() or 0
        cat_rows = db.session.execute(db.text(
            "SELECT i.category, COUNT(DISTINCT i.series_number) as series_count, "
            "COUNT(*) as item_count, COALESCE(SUM(i.qty),0) as total_qty "
            "FROM asset.items i WHERE i.is_deleted=false "
            "GROUP BY i.category ORDER BY item_count DESC"
        )).fetchall()
        return jsonify({"success": True, "data": {
            "total_assets": int(total), "total_series": int(series),
            "active": int(active), "disposed": int(disposed),
            "under_repair": int(under_repair), "total_qty": int(total_qty),
            "by_category": [{"category": r[0], "series_count": int(r[1]),
                             "item_count": int(r[2]), "total_qty": int(r[3])} for r in cat_rows]
        }})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


# ── MODULE USERS ─────────────────────────────────────────────
import json as _json

@asset_bp.route("/users", methods=["GET"])
def get_asset_users():
    try:
        rows = db.session.execute(db.text(
            "SELECT ma.id, ma.user_id, ma.role, ma.permissions, ma.is_active, ma.created_at, "
            "u.email, u.first_name, u.last_name "
            "FROM iam.module_access ma JOIN iam.users u ON ma.user_id = u.id "
            "WHERE ma.module IN ('Asset Management', 'asset') "
            "ORDER BY ma.created_at DESC"
        )).fetchall()
        return jsonify({"success": True, "data": [{
            "id": r[0], "user_id": r[1], "role": r[2],
            "permissions": r[3] if isinstance(r[3], dict) else _json.loads(r[3] or '{}'),
            "is_active": r[4], "created_at": str(r[5]) if r[5] else None,
            "email": r[6], "first_name": r[7] or '', "last_name": r[8] or ''
        } for r in rows]})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e), "data": []})


@asset_bp.route("/users", methods=["POST"])
def add_asset_user():
    data = request.get_json() or {}
    user_id = data.get("user_id")
    role = data.get("role", "viewer")
    permissions = data.get("permissions", {})
    if not user_id:
        return jsonify({"success": False, "message": "user_id required"}), 400
    user = db.session.execute(db.text(
        "SELECT id, email FROM iam.users WHERE id=:id AND is_deleted=false"
    ), {"id": user_id}).first()
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404
    existing = db.session.execute(db.text(
        "SELECT id FROM iam.module_access WHERE user_id=:uid AND module IN ('Asset Management','asset')"
    ), {"uid": user_id}).first()
    if existing:
        return jsonify({"success": False, "message": "User already has access"}), 409
    access_id = str(uuid.uuid4())
    db.session.execute(db.text(
        "INSERT INTO iam.module_access (id, user_id, module, role, permissions, granted_by, tenant_id) "
        "VALUES (:id, :uid, 'Asset Management', :role, :perms, 'system', :tid)"
    ), {"id": access_id, "uid": user_id, "role": role,
        "perms": _json.dumps(permissions), "tid": TENANT_ID})
    db.session.commit()
    return jsonify({"success": True, "message": f"Access granted to {user[1]}"}), 201


@asset_bp.route("/users/<access_id>", methods=["PUT"])
def update_asset_user(access_id):
    data = request.get_json() or {}
    updates, params = [], {"id": access_id}
    if "role" in data:
        updates.append("role=:role"); params["role"] = data["role"]
    if "permissions" in data:
        updates.append("permissions=:permissions"); params["permissions"] = _json.dumps(data["permissions"])
    if not updates:
        return jsonify({"success": False, "message": "Nothing to update"}), 400
    updates.append("updated_at=NOW()")
    db.session.execute(db.text(f"UPDATE iam.module_access SET {', '.join(updates)} WHERE id=:id"), params)
    db.session.commit()
    return jsonify({"success": True, "message": "Permissions updated"})


@asset_bp.route("/users/<access_id>", methods=["DELETE"])
def revoke_asset_user(access_id):
    db.session.execute(db.text("DELETE FROM iam.module_access WHERE id=:id"), {"id": access_id})
    db.session.commit()
    return jsonify({"success": True, "message": "Access revoked"})


# ── EXPORT to Excel ──────────────────────────────────────

@asset_bp.route("/export", methods=["GET"])
def export_assets():
    tid = _get_tenant()
    series_filter = request.args.get("series", "").strip()
    where = f"WHERE i.is_deleted = false AND {TID_COND.replace('tenant_id', 'i.tenant_id')}"
    params = {"tid": tid}
    if series_filter:
        where += " AND i.series_number = :sn"
        params["sn"] = series_filter

    rows = db.session.execute(db.text(
        f"SELECT i.series_number, i.asset_number, i.category, i.description, i.qty, i.make, i.status, i.location "
        f"FROM asset.items i {where} ORDER BY i.series_number ASC, i.asset_number ASC"
    ), params).fetchall()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asset Register"

        # Header style
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Series", "Asset Number", "Category", "Description", "Qty", "Make", "Status", "Location"]
        col_widths = [10, 16, 22, 40, 8, 20, 12, 20]
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 22

        prev_series = None
        for ri, r in enumerate(rows, 2):
            sn = r[0]
            # Only show series number in first row of each group
            series_val = sn if sn != prev_series else None
            prev_series = sn
            vals = [series_val, r[1], r[2], r[3], r[4], r[5], r[6], r[7]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if ci == 1 and v is not None:
                    cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="asset_register.xlsx")
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
